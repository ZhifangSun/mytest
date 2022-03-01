import openpyxl
import numpy as np
import math

file_admet = openpyxl.load_workbook('ADMET.xlsx')
file_admet_training_sheet = file_admet['training']

# store admet file data
file_admet_data, file_admet_index = [], []

for line_data in file_admet_training_sheet.iter_rows(min_row=2, max_row=1975, min_col=2, max_col=6):
    temp = []
    for c in line_data:
        temp.append(c.value)
        #print(temp)
    file_admet_data.append(temp)
#print(file_admet_data)    #ADMET表
#print(len(file_admet_data))
# admet rules, get fit data index, as row num
for step, item in enumerate(file_admet_data):
    count=0
    for i in range(5):
        if i==0 and item[i] == 1:
            count+=1
        if i == 1 and item[i] == 1:
            count += 1
        if i == 2 and item[i] == 0:
            count += 1
        if i == 3 and item[i] == 1:
            count += 1
        if i == 4 and item[i] == 0:
            count += 1
    if count>=3:
        file_admet_index.append(step)
print(file_admet_index)     #将ADMET里‘1’的个数大于等于3的分子索引记录下来
print(len(file_admet_index))
file_molecular = openpyxl.load_workbook('Molecular_Descriptor.xlsx')
file_molecular_training_sheet = file_molecular['training']

# store molecular file training data and columns names
file_molecular_training_data, file_molecular_columns = [], []

# training data
for line_data in file_molecular_training_sheet.iter_rows(min_row=2, max_row=1975, min_col=2, max_col=730):#Molecular_Descriptor.xlsx
    temp = []
    for c in line_data:
        temp.append(c.value)
    file_molecular_training_data.append(temp)
#print(file_molecular_training_data)#Molecular_Descriptor.xlsx
# columns names
for line_data in file_molecular_training_sheet.iter_rows(min_row=1, max_row=1, min_col=2, max_col=730):
    temp = []
    for c in line_data:
        temp.append(c.value)
    file_molecular_columns.append(temp)
#print(file_molecular_columns)     #将729个自变量名称保存
# variables 20
var_names_20 = []
with open('vars_20.txt', 'r', encoding='utf-8') as f:
    nn = f.readlines()
    for item in nn:
        var_names_20.append(item.split(':')[0])
#print(var_names_20)    #保存var_names_20.txt里的自变量名，var_names_20为问题一里选择出的20个变量
# 20 vars index in columns
var_20_names_index = [file_molecular_columns[0].index(item) for item in var_names_20]    #var_20_names_index保存20个自变量在file_molecular_columns中的索引
#print(var_20_names_index)
file_er = openpyxl.load_workbook('ER_activity.xlsx')
file_er_training_sheet = file_er['training']

# store er file training data
file_er_training_data = []

for line_data in file_er_training_sheet.iter_rows(min_row=2, max_row=1975, min_col=2, max_col=2):
    file_er_training_data.append([c.value for c in line_data])
#print(file_er_training_data)   #ER_activity.xlsx中的IC50_nM
# get data that fit rule admet
fit_admet_data = [file_molecular_training_data[i] for i in file_admet_index]    #将ADMET里‘1’的个数大于等于3的分子对应的所有分子描述符信息保存
fit_admet_er_data = [file_er_training_data[i] for i in file_admet_index]       #将ADMET里‘1’的个数大于等于3的分子对应的IC50_nM保存
#print(fit_admet_er_data)
# fit rule er activity   公式
er_rule = sum([item[0] for item in fit_admet_er_data]) / len(file_admet_index)     #log[（将将ADMET里‘1’的个数大于等于3的分子所有的对应的IC50_nM累加）/（ADMET里‘1’的个数大于等于3的分子个数）]
#print(er_rule)
rules_fit = []
for step, item in enumerate(fit_admet_er_data):
    if item[0] < er_rule:
        rules_fit.append(step)
print(len(rules_fit))    #记录下fit_admet_er_data里大于er_rule的索引
fit_rule_data = [fit_admet_data[i] for i in rules_fit]      #记录下fit_admet_data里对应的fit_admet_er_data大于er_rule的分子描述符信息保存

var_20_data = []
for item in fit_rule_data:
    var_20_data.append([item[i] for i in var_20_names_index])    #将var_20_names_index中对应的问题一选出的20的分子描述符对应的描述符信息选出组成632*20的矩阵
#print(var_20_data)
# write to file
var_20_data = np.array(var_20_data)
#print(var_20_data)
f = open('vars_20_value_fit.txt', 'a', encoding='utf-8')
for i in range(20):
    f.write(var_names_20[i]+'\t')
    f.write(str(min(var_20_data[:, i].tolist()))+'  ')
    f.write(str(max(var_20_data[:, i].tolist()))+'\n')
print('done')